from django.shortcuts import render
import openpyxl
from django.core.files.storage import FileSystemStorage
from findCoords import settings
from . import FindLatLong


def home(request):
    if request.method == 'POST' and request.FILES['myfile']: # on user entered data, get the file and assign it to variable to 'myfile'
        myfile = request.FILES['myfile']
        fs = FileSystemStorage()
        filename = fs.save(myfile.name, myfile)
        uploaded_file_url = fs.url(filename)
        wb = openpyxl.load_workbook(myfile) # creating workbook to work with the uploaded excel
        sheet = wb.active
        for i in range(1, sheet.max_row + 1):
            lat, lng = FindLatLong.find_coordinates(sheet.cell(row=i, column=1).value) # call FindLatLong module to find the latitudes and longitudes of the address in the first column
            sheet.cell(row=i, column=2).value = lat
            sheet.cell(row=i,column=3).value = lng
        wb.save(settings.BASE_DIR+uploaded_file_url.replace("/","\\")) # replace the '/' with '\\' in the path of media storage which has been defined in the settings.py
        return render(request, 'home.html', {
            'uploaded_file_url': uploaded_file_url
        }) # return the link to download the modified excel workbook
    return render(request, 'home.html') # if GET request, render original home page